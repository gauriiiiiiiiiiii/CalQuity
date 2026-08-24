"""Structured operational data: accounts, orders, tickets.

Loaded from the supplied workbook with openpyxl (the set is tiny; no pandas needed).
This module owns three things:
  1. account-scoped accessors (access control lives here, keyed off the Session);
  2. the dataset snapshot time, used for ALL time math (never the wall clock);
  3. pure, unit-tested calculators for cancellation fees, service credits, and SLA
     status, so the LLM never does the arithmetic itself.

Contract-specific overrides (Northstar free cancellation, LumenWorks credit terms,
custom SLA targets) are parsed from the agreement PDFs into CONTRACT_TERMS, each with
a citation back to the source clause. A production system would extract these with the
agent + human verification; here they are encoded once, transparently and with sources.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import openpyxl

from core import config
from core.session import Session

IST = ZoneInfo("Asia/Kolkata")

# --- Contract overrides, parsed from the agreement PDFs (with citations) -------
# Keys map to clauses in docs 05 / 06. "sla" targets are (unit, value):
#   ("clock_min", 15) => 15 wall-clock minutes, 24x7
#   ("business_hours", 4) / ("business_days", 2) => business-time targets
CONTRACT_TERMS: dict[str, dict] = {
    "ACCT-001": {  # Northstar Logistics Enterprise Agreement
        "cancellation_fee_waived": True,
        "cancellation_citation": "Northstar Enterprise Agreement §2 (cancel any BOOKED "
                                 "shipment before pickup, no fee, regardless of age).",
        "credit": None,  # §3: default SOP applies unless stated
        "credit_monthly_cap": 5000,
        "credit_citation": "Northstar Enterprise Agreement §3 (monthly credits capped INR 5,000; "
                           "otherwise current SOP applies).",
        "sla": {"P1": ("clock_min", 15), "P2": ("clock_min", 60), "P3": ("business_hours", 8)},
        "sla_citation": "Northstar Enterprise Agreement §1 (replaces standard SLA targets).",
    },
    "ACCT-002": {  # LumenWorks Service Agreement
        "cancellation_fee_waived": False,
        "cancellation_citation": "LumenWorks Service Agreement §2 (no waiver; use current SOP).",
        "credit": {"type": "fixed", "threshold_hours": 4, "amount": 300},
        "credit_citation": "LumenWorks Service Agreement §3 (fixed INR 300 credit when pickup "
                           ">4h past window, carrier fault, no customer fault; replaces SOP default).",
        "sla": {"P1": ("business_hours", 2), "P2": ("business_hours", 4), "P3": ("business_days", 2)},
        "sla_citation": "LumenWorks Service Agreement §1 (no weekend/after-hours coverage).",
    },
}

# --- Default policy parameters (from SOP v4 / Support Policy v3) ---------------
SOP_CANCELLATION_FEE = 250
SOP_FREE_WINDOW_MIN = 30
SOP_CREDIT_DEFAULT = {"type": "default", "threshold_hours": 2, "cap": 500, "pct": 0.10}
SOP_MANAGER_APPROVAL_ABOVE = 1000
SOP_CANCELLATION_CITATION = "Cancellation & Service Credit SOP v4 §1."
SOP_CREDIT_CITATION = "Cancellation & Service Credit SOP v4 §2."

# Support Policy v3 §3 default first-response targets by plan.
DEFAULT_SLA: dict[str, dict[str, tuple[str, float]]] = {
    "Enterprise": {"P1": ("clock_min", 30), "P2": ("clock_min", 120), "P3": ("business_days", 1)},
    "Growth": {"P1": ("business_hours", 2), "P2": ("business_hours", 4), "P3": ("business_days", 2)},
    "Standard": {"P1": ("business_hours", 4), "P2": ("business_days", 1), "P3": ("business_days", 2)},
}
SLA_POLICY_CITATION = "Support Policy v3 §3 (default first-response targets)."

# Business calendar for approximate business-time SLA math (documented simplification).
BIZ_START_H, BIZ_END_H = 9, 18  # 09:00–18:00 IST, Mon–Fri


# --- Loading -----------------------------------------------------------------
def _as_ist(value) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=IST) if value.tzinfo is None else value.astimezone(IST)
    # string fallback
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt).replace(tzinfo=IST)
        except ValueError:
            continue
    return None


def _rows(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append(dict(zip(header, r)))
    return out


class DataStore:
    def __init__(self):
        wb = openpyxl.load_workbook(config.WORKBOOK, data_only=True)
        self.snapshot = self._read_snapshot(wb["README"])
        self.accounts = {r["account_id"]: r for r in _rows(wb["accounts"])}
        self.orders = {r["order_id"]: self._norm_order(r) for r in _rows(wb["orders"])}
        self.tickets = {r["ticket_id"]: self._norm_ticket(r) for r in _rows(wb["tickets"])}
        self.access_log: list[str] = []

    @staticmethod
    def _read_snapshot(ws) -> datetime:
        for row in ws.iter_rows(values_only=True):
            if row and str(row[0]).strip().lower().startswith("dataset snapshot"):
                raw = str(row[1]).strip()  # "2026-08-16 11:00 Asia/Kolkata"
                parts = raw.rsplit(" ", 1)
                stamp = parts[0]
                dt = datetime.strptime(stamp, "%Y-%m-%d %H:%M")
                return dt.replace(tzinfo=IST)
        raise ValueError("Snapshot time not found in README sheet.")

    @staticmethod
    def _norm_order(r: dict) -> dict:
        for f in ("booked_at", "pickup_window_start", "pickup_window_end",
                  "pickup_actual_at", "cancellation_requested_at"):
            r[f] = _as_ist(r.get(f))
        r["carrier_fault"] = bool(r.get("carrier_fault"))
        r["customer_fault"] = bool(r.get("customer_fault"))
        return r

    @staticmethod
    def _norm_ticket(r: dict) -> dict:
        for f in ("created_at", "last_customer_message_at"):
            r[f] = _as_ist(r.get(f))
        return r

    # --- Access-scoped accessors --------------------------------------------
    def _deny(self, session: Session, what: str) -> dict:
        self.access_log.append(f"DENIED {session.label()} -> {what}")
        return {"error": "access_denied",
                "message": f"You are not authorised to access {what}."}

    def get_account(self, session: Session, account_id: str) -> dict:
        if not session.can_access_account(account_id):
            return self._deny(session, f"account {account_id}")
        acct = self.accounts.get(account_id)
        return acct or {"error": "not_found", "message": f"No account {account_id}."}

    def get_order(self, session: Session, order_id: str) -> dict:
        o = self.orders.get(order_id)
        if not o:
            return {"error": "not_found", "message": f"No order {order_id}."}
        if not session.can_access_account(o["account_id"]):
            return self._deny(session, f"order {order_id}")
        if session.is_ops:
            self.access_log.append(f"OPS {session.label()} read order {order_id}")
        return o

    def get_ticket(self, session: Session, ticket_id: str) -> dict:
        t = self.tickets.get(ticket_id)
        if not t:
            return {"error": "not_found", "message": f"No ticket {ticket_id}."}
        if not session.can_access_account(t["account_id"]):
            return self._deny(session, f"ticket {ticket_id}")
        return t

    def list_orders(self, session: Session, account_id: str | None = None) -> list[dict]:
        target = self._scope(session, account_id)
        return [o for o in self.orders.values()
                if target is None or o["account_id"] == target]

    def list_tickets(self, session: Session, account_id: str | None = None,
                     include_closed: bool = True) -> list[dict]:
        target = self._scope(session, account_id)
        out = [t for t in self.tickets.values()
               if target is None or t["account_id"] == target]
        if not include_closed:
            out = [t for t in out if str(t.get("status")) != "closed"]
        return out

    def _scope(self, session: Session, account_id: str | None) -> str | None:
        """Resolve the account filter, enforcing customer scoping. Returns the
        account_id to filter by, or None meaning 'all' (ops only)."""
        if session.role.value == "customer":
            return session.account_id  # always forced to own account
        return account_id  # ops: honour the requested filter, or all if None


# --- Pure calculators (tested against docs/ANALYSIS.md ground truth) ----------
def assess_cancellation(*, status: str, booked_at: datetime | None,
                        cancel_requested_at: datetime | None, now: datetime,
                        fee_waived_by_contract: bool, contract_citation: str | None) -> dict:
    status = (status or "").upper()
    cites = [SOP_CANCELLATION_CITATION]
    if status == "DRAFT":
        return {"status": status, "cancellable": True, "fee_inr": 0,
                "rule": "DRAFT orders cancel free.", "citations": cites}
    if status == "PICKED_UP":
        return {"status": status, "cancellable": False, "fee_inr": None,
                "rule": "PICKED_UP: do not cancel; use the return-to-origin workflow.",
                "next_step": "return_to_origin", "citations": cites}
    if status == "DELIVERED":
        return {"status": status, "cancellable": False, "fee_inr": None,
                "rule": "DELIVERED orders cannot be cancelled.", "citations": cites}
    if status == "BOOKED":
        ref = cancel_requested_at or now
        mins = (ref - booked_at).total_seconds() / 60 if booked_at else None
        within = mins is not None and mins <= SOP_FREE_WINDOW_MIN
        base_fee = 0 if within else SOP_CANCELLATION_FEE
        waived = fee_waived_by_contract and not within
        final_fee = 0 if (within or fee_waived_by_contract) else SOP_CANCELLATION_FEE
        if waived and contract_citation:
            cites = [contract_citation, SOP_CANCELLATION_CITATION]
        rule = (f"BOOKED, not picked up. {round(mins) if mins is not None else '?'} min after "
                f"booking; free within {SOP_FREE_WINDOW_MIN} min, else INR {SOP_CANCELLATION_FEE}"
                + (", waived by contract." if waived else "."))
        return {"status": status, "cancellable": True,
                "minutes_since_booking": round(mins) if mins is not None else None,
                "within_free_window": within, "fee_before_contract": base_fee,
                "fee_waived_by_contract": waived, "fee_inr": final_fee,
                "rule": rule, "citations": cites}
    return {"status": status, "cancellable": None,
            "rule": f"Unrecognised status '{status}'.", "citations": cites}


def assess_service_credit(*, carrier_fault: bool | None, customer_fault: bool | None,
                          pickup_window_end: datetime | None, pickup_actual_at: datetime | None,
                          now: datetime, shipment_fee: float, credit_terms: dict | None,
                          credit_citation: str | None) -> dict:
    cites = [SOP_CREDIT_CITATION]
    terms = credit_terms or SOP_CREDIT_DEFAULT
    threshold = terms.get("threshold_hours", 2)
    if pickup_window_end is None:
        return {"eligible": None, "reason": "No scheduled pickup window on record.",
                "citations": cites}
    end_ref = pickup_actual_at or now  # ongoing delay if still not picked up
    delay_hours = (end_ref - pickup_window_end).total_seconds() / 3600
    picked_up = pickup_actual_at is not None

    if carrier_fault is None or customer_fault is None:
        return {"eligible": None, "delay_hours": round(delay_hours, 2),
                "reason": "Fault is unknown; do not promise a credit, verify first.",
                "citations": cites}

    eligible = (carrier_fault and not customer_fault and delay_hours > threshold)
    if terms.get("type") == "fixed":
        credit = float(terms["amount"])
        if credit_citation:
            cites = [credit_citation, SOP_CREDIT_CITATION]
    else:
        credit = min(terms.get("cap", 500), terms.get("pct", 0.10) * shipment_fee)
    needs_approval = eligible and credit > SOP_MANAGER_APPROVAL_ABOVE

    reasons = []
    if not carrier_fault:
        reasons.append("carrier is not at fault")
    if customer_fault:
        reasons.append("customer is at fault")
    if delay_hours <= threshold:
        reasons.append(f"delay {round(delay_hours,2)}h ≤ {threshold}h threshold")
    reason = "Eligible." if eligible else "Not eligible: " + "; ".join(reasons) + "."

    return {"eligible": eligible, "delay_hours": round(delay_hours, 2),
            "threshold_hours": threshold, "picked_up": picked_up,
            "credit_inr": credit if eligible else 0,
            "needs_manager_approval": needs_approval, "reason": reason, "citations": cites}


def _business_minutes(start: datetime, end: datetime) -> float:
    """Approximate business-time (Mon–Fri, 09:00–18:00 IST) between two instants."""
    if end <= start:
        return 0.0
    total = 0.0
    cur = start
    step = timedelta(minutes=5)
    while cur < end:
        if cur.weekday() < 5 and BIZ_START_H <= cur.hour < BIZ_END_H:
            total += step.total_seconds() / 60
        cur += step
    return total


def assess_sla(*, created_at: datetime, now: datetime, severity: str,
               plan: str, contract_sla: dict | None) -> dict:
    sev = severity.upper()
    if contract_sla and sev in contract_sla:
        unit, val = contract_sla[sev]
        citation = "customer agreement SLA"
        source = "contract"
    else:
        table = DEFAULT_SLA.get(plan, DEFAULT_SLA["Standard"])
        unit, val = table.get(sev, ("business_days", 2))
        citation = SLA_POLICY_CITATION
        source = "policy_v3"

    if unit == "clock_min":
        elapsed = (now - created_at).total_seconds() / 60
        target_min = val
        approximate = False
    elif unit == "business_hours":
        elapsed = _business_minutes(created_at, now)
        target_min = val * 60
        approximate = True
    else:  # business_days => treat 1 business day as 9 business hours
        elapsed = _business_minutes(created_at, now)
        target_min = val * 9 * 60
        approximate = True

    breached = elapsed > target_min
    return {"severity": sev, "sla_source": source, "target_unit": unit, "target_value": val,
            "elapsed_min": round(elapsed), "target_min": round(target_min),
            "breached": breached, "approximate": approximate,
            "remaining_min": round(target_min - elapsed),
            "citations": [citation]}


_STORE: DataStore | None = None


def get_store() -> DataStore:
    global _STORE
    if _STORE is None:
        _STORE = DataStore()
    return _STORE
